from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from extensions import db
from models import Shift, Location, Employee, Record, Attendance
from forms import ShiftForm, LocationForm, EmployeeForm, AttendanceRemarkForm
from utils.excel_parser import generate_template, parse_attendance_file, save_records
from utils.attendance_calc import recalc_all
from utils.excel_exporter import build_export, ATTENDANCE_COLUMNS, ANOMALY_COLUMNS, STATISTICS_COLUMNS
from datetime import datetime, date, timedelta
import os, tempfile

main_bp = Blueprint('main', __name__)
today = date.today()

# ─── Context ────────────────────────────────────────────────
@main_bp.context_processor
def inject_today():
    return {'today': today}

# ─── Helpers ────────────────────────────────────────────────
def get_month_dates(year, month):
    import calendar
    _, last_day = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, last_day + 1)]

def get_month_list(n=6):
    """生成最近 n 个月的列表 yyyy-mm"""
    months = []
    y, m = today.year, today.month
    for _ in range(n):
        months.append(f'{y}-{m:02d}')
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return months

STATUS_CN = {
    'normal': '正常',
    'late': '迟到',
    'early_leave': '早退',
    'both': '迟到+早退',
    'no_punch_in': '缺少上班卡',
    'no_punch_out': '缺少下班卡',
    'absent': '缺勤',
    'anomaly': '异常',
}

# ─── Dashboard ──────────────────────────────────────────────
@main_bp.route('/')
def index():
    employees_count = Employee.query.filter_by(is_active=True).count()
    month_str = today.strftime('%Y-%m')

    month_att = Attendance.query.filter(
        Attendance.date.startswith(today.strftime('%Y-%m'))
    ).all()

    total = len(month_att)
    normal = sum(1 for a in month_att if a.status == 'normal')
    late = sum(1 for a in month_att if a.status in ('late', 'both'))
    anomaly = sum(1 for a in month_att if a.is_anomaly and not a.clerk_remark)

    # 最近7天异常
    week_ago = today - timedelta(days=7)
    recent = Attendance.query.filter(
        Attendance.date >= week_ago,
        Attendance.is_anomaly == True,
    ).order_by(Attendance.date.desc()).limit(10).all()

    stats = {
        'total_employees': employees_count,
        'attendance_rate': round(normal / total * 100, 1) if total else 0,
        'late_count': late,
        'anomaly_count': anomaly,
    }

    return render_template('index.html', stats=stats, recent_anomalies=recent)

# ─── Shifts ─────────────────────────────────────────────────
@main_bp.route('/shifts')
def shifts():
    form = ShiftForm()
    shifts = Shift.query.order_by(Shift.id).all()
    return render_template('shifts/list.html', form=form, shifts=shifts)

@main_bp.route('/shifts/<int:id>/data')
def shift_data(id):
    s = Shift.query.get_or_404(id)
    return jsonify({
        'id': s.id,
        'name': s.name,
        'shift_start': s.shift_start_str(),
        'shift_end': s.shift_end_str(),
        'late_minutes': s.late_minutes,
        'early_leave_minutes': s.early_leave_minutes,
    })

@main_bp.route('/shifts/save', methods=['POST'])
def shift_save():
    form = ShiftForm()
    edit_id = request.form.get('edit_id', type=int)
    if edit_id:
        s = Shift.query.get_or_404(edit_id)
    else:
        s = Shift()

    s.name = form.name.data
    try:
        s.shift_start = datetime.strptime(form.shift_start.data, '%H:%M').time()
        s.shift_end = datetime.strptime(form.shift_end.data, '%H:%M').time()
    except (ValueError, TypeError):
        flash('时间格式无效，请使用 HH:MM 格式', 'error')
        return redirect(url_for('main.shifts'))
    s.late_minutes = form.late_minutes.data or 15
    s.early_leave_minutes = form.early_leave_minutes.data or 15

    if edit_id:
        db.session.commit()
        flash('班次已更新', 'success')
    else:
        db.session.add(s)
        db.session.commit()
        flash('班次已新增', 'success')
    return redirect(url_for('main.shifts'))

@main_bp.route('/shifts/<int:id>/delete', methods=['POST'])
def shift_delete(id):
    s = Shift.query.get_or_404(id)
    s.is_active = not s.is_active
    db.session.commit()
    flash(f'班次已{"启用" if s.is_active else "停用"}', 'success')
    return redirect(url_for('main.shifts'))

# ─── Locations ──────────────────────────────────────────────
@main_bp.route('/locations')
def locations():
    form = LocationForm()
    locs = Location.query.order_by(Location.id).all()
    return render_template('locations/list.html', form=form, locations=locs)

@main_bp.route('/locations/<int:id>/data')
def location_data(id):
    loc = Location.query.get_or_404(id)
    return jsonify({'id': loc.id, 'name': loc.name, 'keywords': loc.keywords})

@main_bp.route('/locations/save', methods=['POST'])
def location_save():
    form = LocationForm()
    edit_id = request.form.get('edit_id', type=int)
    if edit_id:
        loc = Location.query.get_or_404(edit_id)
    else:
        loc = Location()

    loc.name = form.name.data
    loc.keywords = form.keywords.data or ''

    if edit_id:
        db.session.commit()
        flash('地点已更新', 'success')
    else:
        db.session.add(loc)
        db.session.commit()
        flash('地点已新增', 'success')
    return redirect(url_for('main.locations'))

@main_bp.route('/locations/<int:id>/delete', methods=['POST'])
def location_delete(id):
    loc = Location.query.get_or_404(id)
    loc.is_active = not loc.is_active
    db.session.commit()
    flash(f'地点已{"启用" if loc.is_active else "停用"}', 'success')
    return redirect(url_for('main.locations'))

# ─── Employees ──────────────────────────────────────────────
@main_bp.route('/employees')
def employees():
    form = EmployeeForm()
    form.shift_id.choices = [(0, '无')] + [(s.id, s.name) for s in Shift.query.filter_by(is_active=True).all()]
    emps = Employee.query.order_by(Employee.employee_id).all()
    return render_template('employees/list.html', form=form, employees=emps)

@main_bp.route('/employees/<int:id>/data')
def employee_data(id):
    e = Employee.query.get_or_404(id)
    return jsonify({
        'id': e.id,
        'employee_id': e.employee_id,
        'name': e.name,
        'shift_id': e.shift_id or '',
    })

@main_bp.route('/employees/save', methods=['POST'])
def employee_save():
    form = EmployeeForm()
    edit_id = request.form.get('edit_id', type=int)
    if edit_id:
        e = Employee.query.get_or_404(edit_id)
    else:
        e = Employee()

    e.employee_id = form.employee_id.data
    e.name = form.name.data
    e.shift_id = form.shift_id.data if form.shift_id.data and form.shift_id.data != 0 else None

    try:
        if edit_id:
            db.session.commit()
            flash('员工已更新', 'success')
        else:
            db.session.add(e)
            db.session.commit()
            flash('员工已新增', 'success')
    except Exception as ex:
        db.session.rollback()
        flash(f'保存失败：工号可能重复 ({ex})', 'error')
    return redirect(url_for('main.employees'))

@main_bp.route('/employees/<int:id>/delete', methods=['POST'])
def employee_delete(id):
    e = Employee.query.get_or_404(id)
    e.is_active = not e.is_active
    db.session.commit()
    flash(f'员工已{"启用" if e.is_active else "停用"}', 'success')
    return redirect(url_for('main.employees'))

@main_bp.route('/employees/import', methods=['POST'])
def employee_import():
    f = request.files.get('file')
    if not f:
        flash('请选择文件', 'error')
        return redirect(url_for('main.employees'))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    f.save(tmp.name)
    tmp.close()

    from openpyxl import load_workbook
    wb = load_workbook(tmp.name)
    ws = wb.active
    imported = 0
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:
            continue
        emp_id = str(row[0]).strip()
        name = str(row[1]).strip()
        if Employee.query.filter_by(employee_id=emp_id).first():
            errors.append(f'第{i}行: 工号 {emp_id} 已存在')
            continue
        db.session.add(Employee(employee_id=emp_id, name=name))
        imported += 1
    db.session.commit()
    os.unlink(tmp.name)
    flash(f'导入完成：成功 {imported} 条，跳过 {len(errors)} 条', 'success')
    return redirect(url_for('main.employees'))

@main_bp.route('/employees/template')
def download_employee_template():
    from openpyxl import Workbook
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.append(['工号', '姓名'])
    ws.append(['EMP001', '张三'])
    ws.append(['EMP002', '李四'])
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True, download_name='员工导入模板.xlsx')

# ─── Import Data ────────────────────────────────────────────
@main_bp.route('/import')
def import_data():
    return render_template('import_data/list.html')

@main_bp.route('/import/template')
def download_template():
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    generate_template(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name='考勤数据导入模板.xlsx')

@main_bp.route('/import/upload', methods=['POST'])
def import_upload():
    f = request.files.get('file')
    if not f:
        flash('请选择文件', 'error')
        return redirect(url_for('main.import_data'))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    f.save(tmp.name)
    tmp.close()

    records, errors = parse_attendance_file(tmp.name)
    if not records and not errors:
        os.unlink(tmp.name)
        flash('文件中无有效数据', 'error')
        return redirect(url_for('main.import_data'))

    batch, saved = save_records(records)
    att_results = recalc_all(batch=batch)
    att_count = len(att_results)
    anomaly_count = sum(1 for a in att_results if a.is_anomaly)

    os.unlink(tmp.name)

    return render_template('import_data/list.html', result={
        'saved': saved,
        'att_count': att_count,
        'anomaly_count': anomaly_count,
        'errors': errors,
    })

# ─── Attendance ─────────────────────────────────────────────
@main_bp.route('/attendance')
def attendance_view():
    month_str = request.args.get('month', today.strftime('%Y-%m'))
    query = request.args.get('q', '').strip()

    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        year, month = today.year, today.month

    days = get_month_dates(year, month)

    # 获取所有有打卡记录的员工
    emps = Employee.query.filter_by(is_active=True).order_by(Employee.employee_id).all()

    table_data = []
    for emp in emps:
        if query and query not in emp.name and query not in emp.employee_id:
            continue

        att_map = {}
        atts = Attendance.query.filter(
            Attendance.employee_id == emp.employee_id,
            Attendance.date.between(days[0], days[-1])
        ).all()
        for a in atts:
            att_map[a.date.isoformat()] = a

        row = []
        for d in days:
            ds = d.isoformat()
            if ds in att_map:
                a = att_map[ds]
                row.append({'status': a.status})
            else:
                row.append({'status': 'absent'})

        table_data.append((emp.employee_id, emp.name, row))

    months = get_month_list()
    all_columns = ATTENDANCE_COLUMNS
    return render_template('attendance/list.html',
                           table_data=table_data, days=[d.day for d in days],
                           months=months, current_month=month_str, query=query,
                           all_columns=all_columns, STATUS_CN=STATUS_CN)

# ─── Anomalies ──────────────────────────────────────────────
@main_bp.route('/anomalies')
def anomalies():
    form = AttendanceRemarkForm()
    filter_ = request.args.get('filter', 'all')

    q = Attendance.query.filter(Attendance.is_anomaly == True).order_by(Attendance.date.desc())

    if filter_ == 'unresolved':
        q = q.filter((Attendance.clerk_remark == '') | (Attendance.clerk_remark.is_(None)))
    elif filter_ == 'resolved':
        q = q.filter(Attendance.clerk_remark != '', Attendance.clerk_remark.isnot(None))

    return render_template('anomalies/list.html',
                           anomalies=q.all(), form=form, filter=filter_,
                           all_columns=ANOMALY_COLUMNS, STATUS_CN=STATUS_CN)

@main_bp.route('/anomalies/<int:id>/data')
def anomaly_data(id):
    a = Attendance.query.get_or_404(id)
    return jsonify({'id': a.id, 'remark': a.clerk_remark or ''})

@main_bp.route('/anomalies/<int:id>/remark', methods=['POST'])
def anomaly_remark(id):
    a = Attendance.query.get_or_404(id)
    a.clerk_remark = request.form.get('clerk_remark', '')
    db.session.commit()
    flash('备注已保存', 'success')
    return redirect(url_for('main.anomalies'))

# ─── Statistics ─────────────────────────────────────────────
@main_bp.route('/statistics')
def statistics():
    month_str = request.args.get('month', today.strftime('%Y-%m'))
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        year, month = today.year, today.month

    days = get_month_dates(year, month)
    months = get_month_list()

    atts = Attendance.query.filter(
        Attendance.date.between(days[0], days[-1])
    ).all()

    # 汇总
    normal = sum(1 for a in atts if a.status == 'normal')
    late = sum(1 for a in atts if a.status in ('late', 'both'))
    early_leave = sum(1 for a in atts if a.status == 'early_leave')
    missed = sum(1 for a in atts if a.status in ('no_punch_in', 'no_punch_out', 'absent'))
    anomaly = sum(1 for a in atts if a.is_anomaly)

    summary = {
        'total_days': len(days),
        'normal': normal,
        'late': late,
        'early_leave': early_leave,
        'missed': missed,
        'anomaly': anomaly,
    }

    # 个人排行
    stats = {}
    for a in atts:
        if a.employee_id not in stats:
            stats[a.employee_id] = {
                'name': a.employee_name,
                'employee_id': a.employee_id,
                'late': 0, 'early_leave': 0, 'missed': 0
            }
        s = stats[a.employee_id]
        if a.status in ('late', 'both'):
            s['late'] += 1
        if a.status == 'early_leave':
            s['early_leave'] += 1
        if a.status in ('no_punch_in', 'no_punch_out', 'absent'):
            s['missed'] += 1

    stats_list = sorted(stats.values(), key=lambda x: -(x['late'] + x['missed']))

    import json

    # Chart.js 数据（每日正常 vs 异常）
    chart_labels = []
    chart_normal = []
    chart_late = []
    chart_early = []
    chart_missed = []

    for d in days:
        ds = d.strftime('%m/%d')
        day_atts = [a for a in atts if a.date == d]
        chart_labels.append(ds)
        chart_normal.append(sum(1 for a in day_atts if a.status == 'normal'))
        chart_late.append(sum(1 for a in day_atts if a.status in ('late', 'both')))
        chart_early.append(sum(1 for a in day_atts if a.status == 'early_leave'))
        chart_missed.append(sum(1 for a in day_atts if a.status in ('no_punch_in', 'no_punch_out', 'absent')))

    # 饼图
    pie = {}
    for a in atts:
        if a.is_anomaly:
            cn = STATUS_CN.get(a.status, a.status)
            pie[cn] = pie.get(cn, 0) + 1

    chart_data = {
        'labels': json.dumps(chart_labels, ensure_ascii=False),
        'normal': json.dumps(chart_normal),
        'late': json.dumps(chart_late),
        'early_leave': json.dumps(chart_early),
        'missed': json.dumps(chart_missed),
        'anomaly_labels': json.dumps(list(pie.keys()), ensure_ascii=False),
        'anomaly_values': json.dumps(list(pie.values())),
    }

    return render_template('statistics/list.html', stats=stats_list,
                           summary=summary, chart_data=chart_data,
                           months=months, current_month=month_str,
                           all_columns=STATISTICS_COLUMNS)


# ─── Export ────────────────────────────────────────────────
@main_bp.route('/attendance/export', methods=['POST'])
def attendance_export():
    columns_json = request.form.get('columns', '[]')
    import json
    selected = json.loads(columns_json)

    month_str = request.form.get('month', today.strftime('%Y-%m'))
    query = request.form.get('q', '').strip()
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        year, month = today.year, today.month

    days = get_month_dates(year, month)
    col_defs = [(k, n) for k, n in ATTENDANCE_COLUMNS if k in selected or not selected]
    if not col_defs:
        col_defs = ATTENDANCE_COLUMNS

    emps = Employee.query.filter_by(is_active=True).order_by(Employee.employee_id).all()
    atts = Attendance.query.filter(
        Attendance.date.between(days[0], days[-1])
    ).order_by(Attendance.employee_id, Attendance.date).all()

    att_map = {}
    for a in atts:
        att_map[(a.employee_id, a.date.isoformat())] = a

    rows = []
    for emp in emps:
        if query and query not in emp.name and query not in emp.employee_id:
            continue
        for d in days:
            ds = d.isoformat()
            a = att_map.get((emp.employee_id, ds))
            if a and a.status != 'absent':
                rows.append(build_att_row(a, emp))
            else:
                rows.append({'employee_name': emp.name, 'employee_id': emp.employee_id,
                             'date': str(d), 'status_cn': '缺勤', 'punch_in': '', 'punch_out': '',
                             'shift_name': '', 'location_name': '', 'clerk_remark': ''})

    buf = build_export(col_defs, rows, f'考勤_{month_str}')
    return send_file(buf, as_attachment=True,
                     download_name=f'考勤数据_{month_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main_bp.route('/anomalies/export', methods=['POST'])
def anomalies_export():
    columns_json = request.form.get('columns', '[]')
    import json
    selected = json.loads(columns_json)

    filter_ = request.form.get('filter', 'all')
    q = Attendance.query.filter(Attendance.is_anomaly == True).order_by(Attendance.date.desc())
    if filter_ == 'unresolved':
        q = q.filter((Attendance.clerk_remark == '') | (Attendance.clerk_remark.is_(None)))
    elif filter_ == 'resolved':
        q = q.filter(Attendance.clerk_remark != '', Attendance.clerk_remark.isnot(None))

    col_defs = [(k, n) for k, n in ANOMALY_COLUMNS if k in selected or not selected]
    if not col_defs:
        col_defs = ANOMALY_COLUMNS

    rows = [build_att_row(a) for a in q.all()]
    buf = build_export(col_defs, rows, '异常反馈')
    return send_file(buf, as_attachment=True,
                     download_name='异常反馈.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main_bp.route('/statistics/export', methods=['POST'])
def statistics_export():
    columns_json = request.form.get('columns', '[]')
    import json
    selected = json.loads(columns_json)

    month_str = request.form.get('month', today.strftime('%Y-%m'))
    try:
        year, month = map(int, month_str.split('-'))
    except ValueError:
        year, month = today.year, today.month

    days = get_month_dates(year, month)
    atts = Attendance.query.filter(Attendance.date.between(days[0], days[-1])).all()

    col_defs = [(k, n) for k, n in STATISTICS_COLUMNS if k in selected or not selected]
    if not col_defs:
        col_defs = STATISTICS_COLUMNS

    stats = {}
    for a in atts:
        if a.employee_id not in stats:
            stats[a.employee_id] = {
                'employee_name': a.employee_name,
                'employee_id': a.employee_id,
                'late': 0, 'early_leave': 0, 'missed': 0, 'normal': 0, 'total': 0
            }
        s = stats[a.employee_id]
        s['total'] += 1
        if a.status in ('late', 'both'):
            s['late'] += 1
        elif a.status == 'early_leave':
            s['early_leave'] += 1
        elif a.status in ('no_punch_in', 'no_punch_out', 'absent'):
            s['missed'] += 1
        else:
            s['normal'] += 1

    buf = build_export(col_defs, list(stats.values()), f'统计_{month_str}')
    return send_file(buf, as_attachment=True,
                     download_name=f'考勤统计_{month_str}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def build_att_row(a, emp=None):
    """将 Attendance 记录转为导出行"""
    from models import Shift, Location
    shift_name = ''
    if a.shift_id:
        shift = Shift.query.get(a.shift_id)
        if shift:
            shift_name = shift.name
    loc_name = ''
    if a.location_id:
        loc = Location.query.get(a.location_id)
        if loc:
            loc_name = loc.name
    return {
        'employee_name': a.employee_name,
        'employee_id': a.employee_id,
        'date': str(a.date),
        'shift_name': shift_name,
        'punch_in': a.punch_in.strftime('%H:%M') if a.punch_in else '',
        'punch_out': a.punch_out.strftime('%H:%M') if a.punch_out else '',
        'location_name': loc_name,
        'status_cn': STATUS_CN.get(a.status, a.status),
        'clerk_remark': a.clerk_remark or '',
    }
