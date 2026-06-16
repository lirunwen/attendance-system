import openpyxl
from datetime import datetime, date, time
from models import Location, Record, Employee
from extensions import db
import uuid

def generate_template(filepath):
    """生成 Excel 导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '打卡数据'
    headers = ['姓名', '工号', '日期', '打卡时间', '打卡地点']
    ws.append(headers)
    ws.append(['张三', 'EMP001', '2026-06-01', '2026-06-01 09:00:00', '主楼'])
    ws.append(['张三', 'EMP001', '2026-06-01', '2026-06-01 18:00:00', '主楼'])
    ws.append(['李四', 'EMP002', '2026-06-01', '2026-06-01 08:55:00', '北门'])
    wb.save(filepath)


def parse_attendance_file(filepath):
    """解析上传的 Excel 打卡数据，返回 (records, errors)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    header_map = {}
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if cell:
            header_map[cell.strip()] = i

    required = ['姓名', '工号', '日期', '打卡时间']
    for col in required:
        if col not in header_map:
            return [], [f'缺少必需列: {col}']

    records = []
    errors = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if all(v is None for v in row):
            continue

        name = str(row[header_map['姓名']] or '').strip()
        emp_id = str(row[header_map['工号']] or '').strip()

        date_val = row[header_map['日期']]
        if isinstance(date_val, datetime):
            date_val = date_val.date()
        elif isinstance(date_val, date):
            pass
        elif isinstance(date_val, str):
            try:
                date_val = datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
            except ValueError:
                try:
                    date_val = datetime.strptime(date_val.strip(), '%Y/%m/%d').date()
                except ValueError:
                    errors.append(f'第{row_num}行: 日期格式无效 "{date_val}"')
                    continue
        else:
            errors.append(f'第{row_num}行: 日期格式无效')
            continue

        time_val = row[header_map['打卡时间']]
        if isinstance(time_val, datetime):
            punch_time = time_val
        elif isinstance(time_val, str):
            try:
                punch_time = datetime.strptime(time_val.strip(), '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    punch_time = datetime.strptime(time_val.strip(), '%Y-%m-%d %H:%M')
                except ValueError:
                    errors.append(f'第{row_num}行: 打卡时间格式无效 "{time_val}"')
                    continue
        else:
            errors.append(f'第{row_num}行: 打卡时间格式无效')
            continue

        location_raw = ''
        if '打卡地点' in header_map:
            location_raw = str(row[header_map['打卡地点']] or '').strip()

        records.append({
            'employee_id': emp_id,
            'employee_name': name,
            'date': date_val,
            'punch_time': punch_time,
            'location_raw': location_raw,
        })

    return records, errors


def match_location(location_raw):
    """根据关键词匹配地点"""
    if not location_raw:
        return None
    locations = Location.query.filter_by(is_active=True).all()
    for loc in locations:
        if not loc.keywords:
            continue
        for kw in loc.keywords.split(','):
            kw = kw.strip()
            if kw and kw in location_raw:
                return loc.id
    return None


def save_records(records):
    """保存打卡记录到数据库"""
    batch = uuid.uuid4().hex[:12].upper()
    saved = 0
    for r in records:
        loc_id = match_location(r['location_raw'])
        record = Record(
            employee_id=r['employee_id'],
            employee_name=r['employee_name'],
            date=r['date'],
            punch_time=r['punch_time'],
            location_raw=r['location_raw'],
            location_id=loc_id,
            import_batch=batch,
        )
        db.session.add(record)
        saved += 1
    db.session.commit()
    return batch, saved
