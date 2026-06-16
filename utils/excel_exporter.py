"""Excel 数据导出工具 — 支持自定义列选择"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

HEADER_FILL = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)
CENTER = Alignment(horizontal='center', vertical='center')


def build_export(columns, rows, sheet_name='Sheet1'):
    """
    生成 Excel 导出文件，返回 BytesIO

    columns: [(field_key, display_name), ...]
    rows: [{field_key: value, ...}, ...]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写表头
    headers = [col[1] for col in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 写数据
    keys = [col[0] for col in columns]
    for row_data in rows:
        ws.append([row_data.get(k, '') for k in keys])

    # 自动列宽
    for i, _ in enumerate(columns, 1):
        col_letter = get_column_letter(i)
        max_len = len(str(ws[1][i - 1].value or ''))
        for row in ws.iter_rows(min_row=2, max_col=i, max_row=ws.max_row):
            for cell in row:
                cell.border = THIN_BORDER
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── 预定义列定义 ──────────────────────────────────────────

ATTENDANCE_COLUMNS = [
    ('employee_name', '姓名'),
    ('employee_id', '工号'),
    ('date', '日期'),
    ('shift_name', '班次'),
    ('punch_in', '上班打卡'),
    ('punch_out', '下班打卡'),
    ('location_name', '打卡地点'),
    ('status_cn', '状态'),
    ('clerk_remark', '文员备注'),
]

ANOMALY_COLUMNS = [
    ('employee_name', '姓名'),
    ('employee_id', '工号'),
    ('date', '日期'),
    ('status_cn', '异常类型'),
    ('punch_in', '上班打卡'),
    ('punch_out', '下班打卡'),
    ('shift_name', '班次'),
    ('location_name', '打卡地点'),
    ('clerk_remark', '文员备注'),
]

STATISTICS_COLUMNS = [
    ('employee_name', '姓名'),
    ('employee_id', '工号'),
    ('late', '迟到次数'),
    ('early_leave', '早退次数'),
    ('missed', '缺卡次数'),
    ('normal', '正常天数'),
    ('total', '总计天数'),
]
